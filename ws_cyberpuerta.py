import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from bs4 import BeautifulSoup
import re
import time

# Configura el WebDriver
s = Service('C:\\Users\\Abima\\Downloads\\edgedriver_win64\\msedgedriver.exe')
driver = webdriver.Edge(service=s)

# Abre la URL
url = 'https://www.cyberpuerta.mx/Audio-y-Video/TV-y-Pantallas/Pantallas/'
driver.get(url)
time.sleep(5)  # Espera 5 segundos
# Obtén el contenido de la página
page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')

# Lista para almacenar la información de los monitores
monitors = []

# Encuentra los elementos que contienen la información de los monitores
monitor_elements = soup.find_all('li', class_=re.compile(r'cell productData small-12 small-order-\d+'))

for monitor_element in monitor_elements:
    print(monitor_element)
    title = monitor_element.select_one('a.emproduct_right_title').text.strip()
    price = monitor_element.find('label', {'class': 'price'}).text.strip()
    resolution = monitor_element.find(string=re.compile('Resolución de la pantalla')).findNext('span').text.strip()
    screen_size = monitor_element.find(string=re.compile('Diagonal de la pantalla')).findNext('span').text.strip()

    # Añade la información del monitor a la lista
    monitors.append({
        'Titulo': title,
        'Precio': price,
        'Resolución': resolution,
        'Tamaño de pantalla': screen_size
    })

# Crear un nuevo libro de trabajo y seleccionar la hoja activa
wb = openpyxl.Workbook()
sheet = wb.active

# Escribir el encabezado en la primera fila
header = ['Titulo', 'Precio', 'Resolución', 'Tamaño de pantalla']
for col_num, header_text in enumerate(header, 1):
    sheet.cell(row=1, column=col_num).value = header_text

# Escribir los datos en las siguientes filas
for row_num, monitor in enumerate(monitors, 2):
    sheet.cell(row=row_num, column=1).value = monitor['Titulo']
    sheet.cell(row=row_num, column=2).value = monitor['Precio']
    sheet.cell(row=row_num, column=3).value = monitor['Resolución']
    sheet.cell(row=row_num, column=4).value = monitor['Tamaño de pantalla']

# Guardar el libro de trabajo en un archivo
wb.save('ws_results\\pantallascyberpuerta.xlsx')

driver.quit()
