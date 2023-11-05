import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from bs4 import BeautifulSoup
import re
import time

# Configura el WebDriver
s = Service('C:\\Users\\Abima\\Downloads\\edgedriver_win64\\msedgedriver.exe')
driver = webdriver.Edge(service=s)

# Abre la URL
url = 'https://intercompras.com/c/tvs-pantallas-1143'
driver.get(url)
time.sleep(2)  # Espera 5 segundos
# Obtén el contenido de la página
page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')

# Lista para almacenar la información de los monitores
monitors = []

# Encuentra los elementos que contienen la información de los monitores
monitor_elements = soup.find_all('div', class_='divContentProductInfo')

for monitor_element in monitor_elements:
    print(monitor_element.prettify())
    title = monitor_element.select_one('a.spanProductListInfoTitle').text.strip()
    price = monitor_element.select_one('div.divProductListPrice').text.strip()
    resolution_div = monitor_element.find('div', text=re.compile('Resolución de la pantalla'))
    resolution = resolution_div.find_next_sibling('div').text.strip() if resolution_div else 'No disponible'

    screen_size_div = monitor_element.find('div', text=re.compile('Tamaño de pantalla'))
    screen_size = screen_size_div.find_next_sibling('div').text.strip() if screen_size_div else 'No disponible'

    # Añade la información del monitor a la lista
    monitors.append({
        'Titulo': title,
        'Precio': price,
        'Resolución': resolution,
        'Tamaño de pantalla': screen_size
    })

# Crear un nuevo libro de trabajo y seleccionar la hoja activa
wb = openpyxl.Workbook()
sheet = wb.active

# Escribir el encabezado en la primera fila
header = ['Titulo', 'Precio', 'Resolución', 'Tamaño de pantalla']
for col_num, header_text in enumerate(header, 1):
    sheet.cell(row=1, column=col_num).value = header_text

# Escribir los datos en las siguientes filas
for row_num, monitor in enumerate(monitors, 2):
    sheet.cell(row=row_num, column=1).value = monitor['Titulo']
    sheet.cell(row=row_num, column=2).value = monitor['Precio']
    sheet.cell(row=row_num, column=3).value = monitor['Resolución']
    sheet.cell(row=row_num, column=4).value = monitor['Tamaño de pantalla']

# Guardar el libro de trabajo en un archivo
wb.save('ws_results\\pantallasintercompras.xlsx')

driver.quit()
