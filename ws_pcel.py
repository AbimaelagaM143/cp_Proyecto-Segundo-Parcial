import re
import openpyxl
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.edge.service import Service

# Configura el WebDriver
s = Service('C:\\Users\\Abima\\Downloads\\edgedriver_win64\\msedgedriver.exe')
driver = webdriver.Edge(service=s)

# Abre la URL
url = 'https://pcel.com/electronica/televisores'
driver.get(url)

# Obtén el contenido de la página
page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')

# Definiendo la URL base y headers
base_url = 'https://pcel.com/electronica/televisores'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'
}

monitors = []
monitor_elements = soup.find_all('div', class_='name')

# Modifica la expresión regular
regex = re.compile(r"(?P<tipo>[\w\s]+)\sde\s(?P<pulgadas>\d+(\.\d+)?\")?.*?Resolución\s*(?P<resolucion>[\d\sx]+)?.*?(?P<ms>\d+\sms)?")

for monitor_element in monitor_elements:
    description = monitor_element.a.text
    match = regex.search(description)
    if match:
        # Nueva línea para extraer el precio
        price_div = monitor_element.find_next('div', class_='price')
        price = price_div.text.strip() if price_div else 'No disponible'

        # Añadir el precio al diccionario
        monitor_info = match.groupdict()
        monitor_info['Precio'] = price

        monitors.append(monitor_info)

# getting the "Next →" HTML element
next_li_element = soup.find('li', class_='next')

# if there is a next page to scrape
while next_li_element is not None:
    next_page_relative_url = next_li_element.find('a', href=True)['href']

    # getting the new page
    page = requests.get(base_url + next_page_relative_url, headers=headers)

    # parsing the new page
    soup = BeautifulSoup(page.text, 'html.parser')

    # scraping logic...
    monitor_elements = soup.find_all('div', class_='name')
    for monitor_element in monitor_elements:
        description = monitor_element.a.text
        match = regex.search(description)
        if match:
            monitors.append(match.groupdict())

    # looking for the "Next →" HTML element in the new page
    next_li_element = soup.find('li', class_='next')

# Crear un nuevo libro de trabajo y seleccionar la hoja activa
wb = openpyxl.Workbook()
sheet = wb.active

# Escribir el encabezado en la primera fila
header = ['Tipo', 'Pulgadas', 'Resolución', 'Precio']  # Añadido 'Precio'
for col_num, header_text in enumerate(header, 1):
    sheet.cell(row=1, column=col_num).value = header_text

# Escribir los datos en las siguientes filas
for row_num, monitor in enumerate(monitors, 2):
    sheet.cell(row=row_num, column=1).value = monitor['tipo']
    sheet.cell(row=row_num, column=2).value = monitor['pulgadas']
    sheet.cell(row=row_num, column=3).value = monitor['resolucion']
    sheet.cell(row=row_num, column=4).value = monitor['Precio']  # Añadido columna para 'Precio'

# Guardar el libro de trabajo en un archivo
wb.save('ws_results\\pantallaspcel.xlsx')

driver.quit()
